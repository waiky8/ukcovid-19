import dash
import dash_bootstrap_components as dbc
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
import dash_table
from dash_table.Format import Format, Scheme, Sign, Symbol
import dash_daq as daq
import plotly.graph_objects as go
import pandas as pd
from openpyxl import load_workbook
import urllib.request
import datetime
from datetime import date

# New data >>>
url_ltla = "https://api.coronavirus.data.gov.uk/v2/data?areaType=ltla&metric=cumCasesByPublishDate&metric=newCasesByPublishDate&metric=newDeaths28DaysByPublishDate&metric=cumDeaths28DaysByPublishDate&format=csv"
url_uk = "https://api.coronavirus.data.gov.uk/v2/data?areaType=overview&metric=cumCasesByPublishDate&metric=newCasesByPublishDate&metric=newDeaths28DaysByPublishDate&metric=cumDeaths28DaysByPublishDate&format=csv"

app = dash.Dash(__name__, external_stylesheets=[dbc.themes.COSMO],
                meta_tags=[{"name": "viewport",
                            "content": "width=device-width, initial-scale=1.0, maximum-scale=1.2, minimum-scale=0.5,"
                            }
                           ]
                )
server = app.server
app.title = "UK Covid-19"

# App parameters
topn = 5  # Show highest n values
chart_h = 360  # height of charts
datatable_rows = 500
textcol = "steelblue"  # text colour
bgcol = "white"  # background colour of charts, table etc.
new_cases_col = "steelblue"
new_deaths_col = "firebrick"
tot_cases_col = "darkgrey"
tot_deaths_col = "dimgrey"
grid_col = "gainsboro"
fontsize = 16
display_new = True
display_total = False

today = date.today().isoformat()
upload_date = today
admin_pass = "xxxxxx"

# Read covid file & interrogate dates
df = pd.read_excel("covid_data.xlsx")  # , engine="openpyxl")

date_min = "2020-08-12"  # df["date"].min()
date_max = df["date"].max()

date_list = [str(d) for d in df["date"].unique()]  # list of all dates uploaded

# Rename columns
df = df.rename(columns=
{
    "areaName": "Local Authority",
    "cumCasesByPublishDate": "Total Cases",
    "cumDeaths28DaysByPublishDate": "Total Deaths",
    "newCasesByPublishDate": "New Cases",
    "newDeaths28DaysByPublishDate": "New Deaths",
}
)

df_tbl = df.copy()

# Layout ----------
app.layout = html.Div(
    [
        dbc.Row(dbc.Col(html.H1("UK Covid-19"), style={"text-align": "center", "font-weight": "bold"})),

        dbc.Row(dbc.Col(html.H3("(by Local Authority)"), style={"text-align": "center", "font-weight": "bold"})),

        html.Br(),

        dbc.Col(
            dbc.Card(
                [
                    html.H3("New Cases", className="card-title"),
                    html.H1(
                        id="new_cases",
                        className="card-value",
                        style={"font-weight": "bold"}
                    )
                ],
                style={"color": "white",
                       "background": new_cases_col,
                       "text-align": "center"
                       }
            )
        ),

        dbc.Col(
            dbc.Card(
                [
                    html.H3("New Deaths", className="card-title"),
                    html.H1(
                        id="new_deaths",
                        className="card-value",
                        style={"font-weight": "bold"}
                    )
                ],
                style={"color": "white",
                       "background": new_deaths_col,
                       "text-align": "center"
                       }
            )
        ),

        dbc.Col(
            dbc.Card(
                [
                    html.H3("Total Cases", className="card-title"),
                    html.H1(
                        id="total_cases",
                        className="card-value",
                        style={"font-weight": "bold"}
                    )
                ],
                style={"color": "white",
                       "background": "darkgrey",
                       "text-align": "center"
                       }
            )
        ),

        dbc.Col(
            dbc.Card(
                [
                    html.H3("Total Deaths", className="card-title"),
                    html.H1(
                        id="total_deaths",
                        className="card-value",
                        style={"font-weight": "bold"}
                    )
                ],
                style={"color": "white",
                       "background": "dimgrey",
                       "text-align": "center"
                       }
            )
        ),

        html.Br(), html.Br(),

        dbc.Row(
            [
                dbc.Col(
                    html.Div(
                        dcc.DatePickerSingle(
                            id="date_picker",
                            date=date_max,
                            display_format="MMM D, YYYY",
                            day_size=50,
                            initial_visible_month=date_max,
                            min_date_allowed=date_min,
                            max_date_allowed=today
                        )
                    )
                ),

                html.Br(), html.Br(),

                dbc.Col(
                    html.Div(dbc.Row(
                        [
                            daq.ToggleSwitch(
                                id="toggle_switch",
                                label="New / Cumulative Data",
                                labelPosition="bottom",
                                color=new_cases_col,
                                size=80,
                                value=False
                            )
                        ]
                    )
                    )
                )
            ], style={"padding": "0px 20px 0px 20px"}
        ),

        html.Br(), html.Br(),

        dbc.Col(dcc.Dropdown(
            id="locauth_drop",
            options=[{"label": i, "value": i} for i in sorted(df["Local Authority"].unique())],
            multi=True,
            placeholder="Select Local Authority",
            style={"font-size": fontsize, "color": "black", "background-color": "white"}
        )
        ),

        html.Br(), html.Br(),

        html.Div(dcc.Graph(id="chart1",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           ),
                 style={"padding": "0px 20px 0px 20px"}
                 ),

        html.Br(), html.Br(),

        html.Div(dcc.Graph(id="chart2",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           ),
                 style={"padding": "0px 20px 0px 20px"}
                 ),

        html.Br(),

        html.Div(dcc.Graph(id="chart3",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           )
                 ),

        html.Br(), html.Br(),

        html.Div(dcc.Graph(id="chart4",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           ),
                 style={"padding": "0px 20px 0px 20px"}
                 ),

        html.Br(), html.Br(),

        html.Div(dcc.Graph(id="chart5",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           ),
                 style={"padding": "0px 20px 0px 20px"}
                 ),

        html.Br(), html.Br(),

        html.Div(dcc.Graph(id="chart6",
                           figure={},
                           config={"displayModeBar": False}  # hide plotly controls
                           ),
                 style={"padding": "0px 20px 0px 20px"}
                 ),

        html.Br(), html.Br(),

        html.Div(
            [
                dash_table.DataTable(
                    id="datatable",
                    # columns=[{"name": i, "id": i} for i in df_tbl],
                    columns=[
                        {
                            "id": "Row",
                            "name": "Row",
                            "type": "numeric"
                        },
                        {
                            "id": "Local Authority",
                            "name": "Local Authority",
                            "type": "text"
                        },
                        {
                            "id": "New Cases",
                            "name": "New Cases",
                            "type": "numeric",
                            "format": Format
                                (
                                precision=0,
                                group=",",
                                scheme=Scheme.fixed,
                                symbol=""
                            )
                        },
                        {
                            "id": "New Deaths",
                            "name": "New Deaths",
                            "type": "numeric",
                            "format": Format
                                (
                                precision=0,
                                group=",",
                                scheme=Scheme.fixed,
                                symbol="",
                            )
                        },
                        {
                            "id": "Total Cases",
                            "name": "Total Cases",
                            "type": "numeric",
                            "format": Format
                                (
                                precision=0,
                                group=",",
                                scheme=Scheme.fixed,
                                symbol=""
                            )
                        },
                        {
                            "id": "Total Deaths",
                            "name": "Total Deaths",
                            "type": "numeric",
                            "format": Format
                                (
                                precision=0,
                                group=",",
                                scheme=Scheme.fixed,
                                symbol=""
                            )
                        }
                    ],

                    sort_action="native",  # native / none
                    sort_mode="single",  # single / multi
                    filter_action="none",  # native / none
                    page_action="native",  # native / none
                    page_current=0,  # current page number
                    page_size=datatable_rows,  # rows per page
                    fill_width=True,

                    style_cell={
                        "height": "30px",
                        "backgroundColor": bgcol,
                        "color": textcol,
                        "font-family": "Verdana",
                        "font_size": fontsize,
                        "padding": "0px 10px 0px 10px"
                    },

                    style_cell_conditional=[
                        {
                            "if": {
                                "column_id": "Row"
                            },
                            "width": "50px"},
                        {
                            "if": {
                                "column_id": "Local Authority"
                            },
                            "width": "80px"},
                        {
                            "if": {
                                "column_id": "New Cases"
                            },
                            "width": "50px",
                            "color": "white",
                            "backgroundColor": new_cases_col},
                        {
                            "if": {
                                "column_id": "New Deaths"
                            },
                            "width": "50px",
                            "color": "white",
                            "backgroundColor": new_deaths_col},
                        {
                            "if": {
                                "column_id": "Total Cases"
                            },
                            "width": "50px",
                            "color": "white",
                            "backgroundColor": tot_cases_col},
                        {
                            "if": {
                                "column_id": "Total Deaths"
                            },
                            "width": "50px",
                            "color": "white",
                            "backgroundColor": tot_deaths_col
                        },
                    ],

                    fixed_rows={"headers": True},
                    fixed_columns={
                        'headers': True,
                        'data': 2
                    },
                    style_data={
                        # wrap long cell content into multiple lines
                        "whiteSpace": "normal",
                        "height": "auto"
                    },
                    style_table={
                        "overflowX": "auto",
                        "overflowY": "auto"
                    },  # "height": "500px",
                    css=[
                        {
                            "selector": ".row",
                            "rule": "margin: 0; flex-wrap: nowrap"
                        }
                    ],  # fix clipping issue
                    style_header={
                        "bold": True,
                        "color": "black",
                        "backgroundColor": "white",
                        "whiteSpace": "normal"
                    },
                    style_as_list_view=True
                )
            ], style={"padding": "0px 20px 0px 20px"}
        ),

        html.Br(), html.Br(),

        dbc.Row(html.H2("Select date for data upload (password required)"),
                style={"padding": "0px 0px 0px 50px"}
                ),

        html.Br(),

        dbc.Col(
            dcc.Input(id="admin_pw", type="password", placeholder="Admin PIN"),
            width={"size": 6, "offset": 1}
        ),

        html.Br(), html.Br(),

        dbc.Col(
            html.Div(
                dcc.DatePickerSingle(
                    id="date_picker2",
                    date=today,
                    display_format="MMM D, YYYY",
                    day_size=50,
                    initial_visible_month=date_max,
                    min_date_allowed=date_min,
                    max_date_allowed=today
                )
            ),
            width={"size": 2, "offset": 1}
        ),

        html.Br(), html.Br(),

        dbc.Col(
            html.Div(id="message", children="", style={"padding": "0px 20px 0px 20px"})
        ),

        html.Br(), html.Br(), html.Br(),

        dbc.Row(html.Label(["Data Source: ",
                            html.A("GovUK", href="https://coronavirus.data.gov.uk/details/download",
                                   target="_blank")]),
                style={"padding": "0px 0px 0px 50px"}
                ),

        dbc.Row(html.Label(["Code: ",
                            html.A("Github", href="https://github.com/waiky8/ukcovid-19", target="_blank")]),
                style={"padding": "0px 0px 0px 50px"}
                ),

        html.Div(id="dummy", children=[], style={"display": "none"})
        # dummy DIV to trigger totals_timeline callback
    ]
)


# Data Table ----------
@app.callback(
    Output("datatable", "data"),
    [
        Input("date_picker", "date"),
        Input("locauth_drop", "value"),
        Input("toggle_switch", "value")
    ]
)
def update_datatable(selected_date, selected_auth, selected_data):
    global df, hide_new, hide_total

    if selected_data == False:
        cases = "New Cases"
        hide_new = True
        hide_total = False
    else:
        cases = "Total Cases"
        hide_new = False
        hide_total = True

    df1 = df[df["date"].isin([selected_date])]

    if (selected_auth is None or selected_auth == []):
        pass
    else:
        df1 = df1[df1["Local Authority"].isin(selected_auth)]

    df1 = df1.sort_values(by=[cases], ascending=False)
    df1["Row"] = df1.reset_index().index
    df1["Row"] += 1

    return df1.to_dict("records")


# Daily Charts ----------
@app.callback(
    [
        Output("chart1", "figure"),
        Output("chart2", "figure")
    ],
    [
        Input("date_picker", "date"),
        Input("locauth_drop", "value"),
        Input("toggle_switch", "value")
    ]
)
def update_graph(selected_date, selected_auth, selected_data):
    global df

    if selected_data == False:
        cases = "New Cases"
        deaths = "New Deaths"
    else:
        cases = "Total Cases"
        deaths = "Total Deaths"

    df1 = df[df["date"].isin([selected_date])]

    if (selected_auth is None or selected_auth == []):
        pass
    else:
        df1 = df1[df1["Local Authority"].isin(selected_auth)]

    data_fig1 = df1.sort_values(by=cases, ascending=False)[:topn]

    fig1 = go.Figure(go.Bar(orientation="h",
                            x=data_fig1[cases],
                            y=data_fig1["Local Authority"],
                            # text=data_fig1["Local Authority"],
                            texttemplate="%{y} - %{x:,}",
                            textposition="inside",
                            insidetextanchor="end"
                            )
                     )

    d = datetime.datetime.strptime(selected_date, "%Y-%m-%d")

    fig1.update_layout(title="<b>" + cases + ": " + d.strftime("%b %d, %Y") + "</b>",
                       title_font_color=textcol,
                       font_color=textcol,
                       font_size=fontsize,
                       showlegend=False,
                       xaxis={"categoryorder": "total descending",
                              "title": "",
                              "tickangle": 0,
                              "visible": False,
                              "showgrid": False,
                              "zeroline": False,
                              "gridcolor": grid_col,
                              "zerolinecolor": grid_col,
                              "fixedrange": True},  # disable zoom & pan
                       yaxis={"title": "",
                              "autorange": "reversed",
                              "visible": False,
                              "showgrid": False,
                              "zeroline": False,
                              "gridcolor": grid_col,
                              "zerolinecolor": grid_col,
                              "fixedrange": True},  # disable zoom & pan
                       hoverlabel=dict(
                           bgcolor="white",
                           font_size=fontsize,
                           font_family="Rockwell"
                       ),
                       height=chart_h,
                       margin=dict(l=0, r=0, t=50, b=0),
                       plot_bgcolor=bgcol,
                       paper_bgcolor=bgcol
                       )

    fig1.update_traces(marker_color=new_cases_col,
                       hovertemplate="%{y}<br>%{x:,}<extra></extra>"
                       )

    data_fig2 = df1.sort_values(by=[deaths], ascending=False)[:topn]
    fig2 = go.Figure(go.Bar(orientation="h",
                            x=data_fig2[deaths],
                            y=data_fig2["Local Authority"],
                            # text=data_fig2["Local Authority"],
                            texttemplate="%{y} - %{x:,}",
                            textposition="inside",
                            insidetextanchor="end"
                            )
                     )

    fig2.update_layout(title="<b>" + deaths + ": " + d.strftime("%b %d, %Y") + "</b>",
                       title_font_color=textcol,
                       font_color=textcol,
                       font_size=fontsize,
                       showlegend=False,
                       xaxis={"categoryorder": "total descending",
                              "title": "",
                              "tickangle": 0,
                              "visible": False,
                              "showgrid": True,
                              "zeroline": False,
                              "gridcolor": grid_col,
                              "zerolinecolor": grid_col,
                              "fixedrange": True},  # disable zoom & pan
                       yaxis={"title": "",
                              "autorange": "reversed",
                              "visible": False,
                              "showgrid": True,
                              "zeroline": False,
                              "gridcolor": grid_col,
                              "zerolinecolor": grid_col,
                              "fixedrange": True},  # disable zoom & pan
                       hoverlabel=dict(
                           bgcolor="white",
                           font_size=fontsize,
                           font_family="Rockwell"
                       ),
                       height=chart_h,
                       margin=dict(l=0, r=0, t=50, b=0),
                       plot_bgcolor=bgcol,
                       paper_bgcolor=bgcol
                       )

    fig2.update_traces(marker_color=new_deaths_col,
                       hovertemplate="%{y}<br>%{x:,}<extra></extra>"
                       )

    return fig1, fig2


# Timeline Charts Set 1 ----------
@app.callback(
    [
        Output("chart3", "figure"),
        Output("chart4", "figure")
    ],
    [
        Input("date_picker", "date"),
        Input("locauth_drop", "value")
    ]
)
def update_graph2(selected_date, selected_auth):
    global df

    if (selected_auth is None or selected_auth == []):
        dfa = df[df["date"].isin([selected_date])]
        dfa = dfa.sort_values(by="New Cases", ascending=False)[:topn]
        locauth_list = dfa["Local Authority"].unique()
        df1 = df[df["Local Authority"].isin(locauth_list)]
    else:
        df1 = df[df["Local Authority"].isin(selected_auth)]
        locauth_list = df1["Local Authority"].unique()

    fig3 = go.Figure()
    fig3.update_layout(
        title="<b>Cases by Local Authority</b>",
        title_font_color=textcol,
        font_color=textcol,
        font_size=fontsize,
        plot_bgcolor=bgcol,
        paper_bgcolor=bgcol,
        height=chart_h,
        margin=dict(l=0, r=0, t=50, b=0),
        xaxis={"categoryorder": "total descending",
               "title": "",
               "tickangle": 0,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        yaxis={"title": "",
               "autorange": True,
               "visible": True,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        hoverlabel=dict(
            # bgcolor="white",
            font_size=fontsize,
            font_family="Rockwell"
        )
    )

    for la in locauth_list:
        dfx = df1[df1["Local Authority"] == la]
        fig3.add_trace(go.Scatter(x=dfx["date"],
                                  y=dfx["New Cases"],
                                  mode="lines",
                                  name=la,
                                  showlegend=True,
                                  hovertemplate="%{y:,}<br>%{x}<extra></extra>"
                                  )
                       )

    fig4 = go.Figure()
    fig4.update_layout(
        title="<b>Deaths by Local Authority</b>",
        title_font_color=textcol,
        font_color=textcol,
        font_size=fontsize,
        plot_bgcolor=bgcol,
        paper_bgcolor=bgcol,
        height=chart_h,
        margin=dict(l=0, r=0, t=50, b=0),
        xaxis={"categoryorder": "total descending",
               "title": "",
               "tickangle": 0,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        yaxis={"title": "",
               "autorange": True,
               "visible": True,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        hoverlabel=dict(
            # bgcolor="white",
            font_size=fontsize,
            font_family="Rockwell"
        )
    )

    for la in locauth_list:
        dfx = df1[df1["Local Authority"] == la]
        fig4.add_trace(go.Scatter(x=dfx["date"],
                                  y=dfx["New Deaths"],
                                  mode="lines",
                                  name=la,
                                  showlegend=True,
                                  hovertemplate="%{y:,}<br>%{x}<extra></extra>"
                                  )
                       )

    return fig3, fig4


# Timeline Charts Set 2 ----------
@app.callback(
    [
        Output("chart5", "figure"),
        Output("chart6", "figure")
    ],
    Input("dummy", "children")
)
def totals_timeline(none):
    date_list = df["date"].unique()

    fig5 = go.Figure()
    fig5.update_layout(
        title="<b>Daily Cases Over Time</b>",
        title_font_color=textcol,
        font_color=textcol,
        font_size=fontsize,
        plot_bgcolor=bgcol,
        paper_bgcolor=bgcol,
        height=chart_h,
        margin=dict(l=0, r=0, t=50, b=0),
        showlegend=False,
        xaxis={"categoryorder": "total descending",
               "title": "",
               "tickangle": 0,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        yaxis={"title": "",
               "autorange": True,
               "visible": True,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        hoverlabel=dict(
            bgcolor="white",
            font_size=fontsize,
            font_family="Rockwell"
        )
    )

    tot_cases = []
    for dt in date_list:
        dfx = df[df["date"] == dt]
        tc = dfx["New Cases"].sum()
        tot_cases.append(tc)

    fig5.add_trace(go.Scatter(x=date_list,
                              y=tot_cases,
                              fill="tonexty",
                              mode="none",
                              name="Cases",
                              showlegend=True,
                              hovertemplate="%{y:,}<br>%{x}<extra></extra>"
                              )
                   )

    fig6 = go.Figure()
    fig6.update_layout(
        title="<b>Daily Deaths Over Time</b>",
        title_font_color=textcol,
        font_size=fontsize,
        font_color=textcol,
        plot_bgcolor=bgcol,
        paper_bgcolor=bgcol,
        height=chart_h,
        margin=dict(l=0, r=0, t=50, b=0),
        showlegend=False,
        xaxis={"categoryorder": "total descending",
               "title": "",
               "tickangle": 0,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        yaxis={"title": "",
               "autorange": True,
               "visible": True,
               "showgrid": True,
               "zeroline": False,
               "gridcolor": grid_col,
               "zerolinecolor": grid_col,
               "fixedrange": True},  # disable zoom & pan
        hoverlabel=dict(
            bgcolor="white",
            font_size=fontsize,
            font_family="Rockwell"
        )
    )

    tot_deaths = []
    for dt in date_list:
        dfx = df[df["date"] == dt]
        td = dfx["New Deaths"].sum()
        tot_deaths.append(td)

    fig6.add_trace(go.Scatter(x=date_list,
                              y=tot_deaths,
                              fill="tonexty",
                              mode="none",
                              name="Deaths",
                              showlegend=True,
                              hovertemplate="%{y:,}<br>%{x}<extra></extra>"
                              )
                   )

    return fig5, fig6


# Summary counts ----------
@app.callback(
    [
        Output("new_cases", "children"),
        Output("new_deaths", "children"),
        Output("total_cases", "children"),
        Output("total_deaths", "children")
    ],
    Input("date_picker", "date")
)
def update_summary(selected_date):
    global df

    # Attempt to read data for selected date
    url = url_uk + "&release=" + selected_date

    try:
        df1 = pd.read_csv(url)
        df1 = df1[df1["date"] == selected_date]

    except urllib.request.HTTPError as e:
        print(e)

    new_cases = format(int(df1["newCasesByPublishDate"]), ",d")
    new_deaths = format(int(df1["newDeaths28DaysByPublishDate"]), ",d")
    total_cases = format(int(df1["cumCasesByPublishDate"]), ",d")
    total_deaths = format(int(df1["cumDeaths28DaysByPublishDate"]), ",d")

    return new_cases, new_deaths, total_cases, total_deaths


# Admin (Upload New Data) ----------
@app.callback(
    Output("message", "children"),
    [
        Input("date_picker2", "date"),
        Input("admin_pw", "value")
    ],
    prevent_initial_call=True
)
def submit_upload(selected_date, password):
    global date_list

    message = ""
    if password == admin_pass:
        pass
    else:
        message = " Please enter valid pin"
        return message

    d = datetime.datetime.strptime(selected_date, "%Y-%m-%d")

    # Check if data already uploaded
    if selected_date in date_list:
        message = d.strftime("%b %d, %Y") + " data already uploaded"
        return message

    # Attempt to read data for selected date
    url = url_ltla + "&release=" + selected_date

    try:
        print("step 1 of 3: read ", url)
        df_load = pd.read_csv(url)
        print("step 2 of 3: extract data for ", d.strftime("%b %d, %Y"))
        df_load = df_load[df_load["date"] == selected_date]

        print("step 3 of 3: write to covid file...")
        writer = pd.ExcelWriter("covid_data.xlsx", engine="openpyxl")
        writer.book = load_workbook("covid_data.xlsx")
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        for sheetname in writer.sheets:
            df_load.to_excel(writer,
                             sheet_name=sheetname,
                             index=False,
                             header=False,
                             startrow=writer.sheets[sheetname].max_row)
            writer.save()

        message = d.strftime("%b %d, %Y") + " data uploaded"
        refresh_data()

    except urllib.request.HTTPError as e:
        message = d.strftime("%b %d, %Y") + " data not yet available"

    except IOError as e:
        message = d.strftime("%b %d, %Y") + "[" + str(e) + "]"

    print(message)

    return message


# Refresh data after each data upload ----------
def refresh_data():
    global df, df_tbl, date_min, date_max, date_list

    df = pd.read_excel("covid_data.xlsx")  # , engine="openpyxl")

    date_min = df["date"].min()
    date_max = df["date"].max()

    date_list = [str(d) for d in df["date"].unique()]  # list of all dates uploaded

    # Rename columns
    df = df.rename(columns=
    {
        "areaName": "Local Authority",
        "cumCasesByPublishDate": "Total Cases",
        "cumDeaths28DaysByPublishDate": "Total Deaths",
        "newCasesByPublishDate": "New Cases",
        "newDeaths28DaysByPublishDate": "New Deaths",
    }
    )

    df_tbl = df.copy()


if __name__ == "__main__":
    app.run_server(debug=True)
