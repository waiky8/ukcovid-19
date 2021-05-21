import dash
import dash_bootstrap_components as dbc
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output
import pandas as pd
from openpyxl import load_workbook
import bs4 as bs
import urllib.request
import datetime
from datetime import date

'''
===========
DATA SOURCE
===========
'''

url_ltla = 'https://api.coronavirus.data.gov.uk/v2/data?areaType=ltla&metric=cumCasesByPublishDate&metric=newCasesByPublishDate&metric=newDeaths28DaysByPublishDate&metric=cumDeaths28DaysByPublishDate&format=csv'
url_uk = 'https://api.coronavirus.data.gov.uk/v2/data?areaType=overview&metric=cumCasesByPublishDate&metric=newCasesByPublishDate&metric=newDeaths28DaysByPublishDate&metric=cumDeaths28DaysByPublishDate&format=csv'

'''
===========
SET-UP DASH
===========
'''

app = dash.Dash(__name__, external_stylesheets=[dbc.themes.COSMO],
                meta_tags=[
                    {'name': 'viewport',
                     'content': 'width=device-width, initial-scale=1.0, maximum-scale=1.2, minimum-scale=0.5,'
                     }
                ]
                )
server = app.server
app.title = 'UK Covid-19'

'''
================
READ EXCEL FILES
================
'''

covid_daily_file = 'covid_data.xlsx'
covid_totals_file = 'covid_totals.xlsx'

df_daily = pd.read_excel(covid_daily_file)
df_totals = pd.read_excel(covid_totals_file)

'''
======================
PARAMETERS & VARIABLES
======================
'''

date_min = '2020-08-12'  # data available from this date
date_max = df_daily['date'].max()
date_today = date.today()

date_list_daily = [str(d) for d in df_daily['date'].unique()]
date_list_totals = [str(d) for d in df_totals['date'].unique()]

# following are missing coordinates from doogal website
coords = [
    ('Aylesbury Vale', 51.8996278, -1.1193516),
    ('Chiltern', 51.6788424, -0.7737162),
    ('South Bucks', 51.5600494, -0.7301907),
    ('Wycombe', 51.6635175, -0.9501978)
]

df_coords = pd.DataFrame(coords, columns=['LA', 'Lat', 'Long'])

'''
===================
DASH LAYOUT SECTION
===================
'''

app.layout = html.Div(
    [
        html.Div(
            [
                html.H1('UK Covid-19'),
                html.H3('(Data Load)')
            ],
            style={'text-align': 'center', 'font-weight': 'bold'}
        ),

        html.Br(),

        html.Div(
            [
                dbc.Row(
                    [
                        dbc.Col(
                            [
                                html.P('Select Date'),

                                dcc.DatePickerSingle(
                                    id='date_picker',
                                    clearable=True,
                                    with_portal=True,
                                    date=date_max,
                                    display_format='MMM D, YYYY',
                                    day_size=50,
                                    initial_visible_month=date_max,
                                    min_date_allowed=date_min,
                                    max_date_allowed=date_today
                                )
                            ]
                        ),

                        dbc.Col(
                            [
                                html.Br(),

                                dcc.Loading(
                                    dbc.Card(
                                        [
                                            html.H5('Covid Data', className='card-title'),
                                            html.H4(
                                                id='message1',
                                                className='card-value',
                                                style={'font-weight': 'bold'}
                                            )
                                        ],
                                        style={
                                            'color': 'white',
                                            'background': 'teal',
                                            'text-align': 'center'
                                        }
                                    )
                                ),

                                html.Br()
                            ]
                        ),

                        dbc.Col(
                            [
                                html.Br(),

                                dcc.Loading(
                                    dbc.Card(
                                        [
                                            html.H5('Covid Totals', className='card-title'),
                                            html.H4(
                                                id='message2',
                                                className='card-value',
                                                style={'font-weight': 'bold'}
                                            )
                                        ],
                                        style={
                                            'color': 'white',
                                            'background': 'midnightblue',
                                            'text-align': 'center'
                                        }
                                    )
                                ),

                                html.Br()
                            ]
                        )
                    ], style={'background': 'whitesmoke', 'border-style': 'groove'}
                )
            ], style={'padding': '0px 30px 0px 30px'}
        )
    ]
)

'''
======================
CALLBACK FOR DATA LOAD
======================
'''


@app.callback(
    [Output('message1', 'children'),
     Output('message2', 'children')
     ],
    Input('date_picker', 'date')
)
def return_new_data(selected_date):
    d = datetime.datetime.strptime(selected_date, '%Y-%m-%d')

    '''
    -----------
    Daily Data
    -----------
    '''

    if selected_date in date_list_daily:
        message1 = 'Already Uploaded 👍'

    else:
        url = url_ltla + '&release=' + selected_date

        try:
            print('Processing daily data...')
            print(str(datetime.datetime.now()), 'step 1 of 3: read ', url)
            df_load = pd.read_csv(url)

            print(str(datetime.datetime.now()), 'step 2 of 3: extract data for ', d.strftime('%b %d, %Y'))
            df_load = df_load[df_load['date'] == selected_date]

            '''
            --------------------
            LATITUDE & LONGITUDE
            --------------------
            '''

            tot_rows = df_load.shape[0]
            latitude = []
            longitude = []

            for i, r in enumerate(df_load.itertuples(), start=1):
                loc_auth = r.areaName

                # if loc_auth == 'Hackney and City of London':
                #     loc_auth = 'Hackney'

                # if loc_auth == 'Cornwall and Isles of Scilly':
                #     loc_auth = 'Cornwall'

                # if loc_auth == 'Comhairle nan Eilean Siar':
                #     loc_auth = 'Na h-Eileanan Siar'

                lat, long = get_coord(loc_auth)

                if lat == '' or long == '':
                    for c in df_coords.itertuples():
                        if c.LA == loc_auth:
                            lat = c.Lat
                            long = c.Long

                if lat == '' or long == '':
                    print('*** Not Found ***')

                latitude.append(lat)
                longitude.append(long)

                print("[", i, "/", tot_rows, "]", loc_auth, lat, long)

            df_load['Latitude'] = latitude
            df_load['Longitude'] = longitude

            '''
            --------------
            WRITE TO EXCEL
            --------------
            '''

            print(str(datetime.datetime.now()), 'step 3 of 3: write to covid file...')
            writer = pd.ExcelWriter(covid_daily_file)
            writer.book = load_workbook(covid_daily_file)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

            for sheetname in writer.sheets:
                df_load.to_excel(
                    writer,
                    sheet_name=sheetname,
                    index=False,
                    header=False,
                    startrow=writer.sheets[sheetname].max_row,
                    columns=
                    [
                        'date',
                        'areaType',
                        'areaCode',
                        'areaName',
                        'cumCasesByPublishDate',
                        'newCasesByPublishDate',
                        'newDeaths28DaysByPublishDate',
                        'cumDeaths28DaysByPublishDate',
                        'Latitude',
                        'Longitude'
                    ]
                )
                writer.save()
                date_list_daily.append(selected_date)

            message1 = 'Upload Complete ✔'

        except urllib.request.HTTPError:
            message1 = 'Not Available ❌'

    '''
    -----------
    Totals Data
    -----------
    '''

    if selected_date in date_list_totals:
        message2 = 'Already Uploaded 👍'

    else:
        url = url_uk + '&release=' + selected_date

        try:
            print('Processing totals data...')
            print(str(datetime.datetime.now()), 'step 1 of 3: read ', url)
            df_load = pd.read_csv(url)

            print(str(datetime.datetime.now()), ' step 2 of 3: extract data for ', d.strftime('%b %d, %Y'))
            df_load = df_load[df_load['date'] == selected_date]

            print(str(datetime.datetime.now()), 'step 3 of 3: write to covid file...')
            writer = pd.ExcelWriter(covid_totals_file)
            writer.book = load_workbook(covid_totals_file)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

            for sheetname in writer.sheets:
                df_load.to_excel(
                    writer,
                    sheet_name=sheetname,
                    index=False,
                    header=False,
                    startrow=writer.sheets[sheetname].max_row,
                    columns=
                    [
                        'date',
                        'areaType',
                        'areaCode',
                        'areaName',
                        'cumCasesByPublishDate',
                        'newCasesByPublishDate',
                        'newDeaths28DaysByPublishDate',
                        'cumDeaths28DaysByPublishDate'
                    ]
                )
                writer.save()
                date_list_totals.append(selected_date)

            message2 = 'Upload Complete ✔'

        except urllib.request.HTTPError:
            message2 = 'Not Available ❌'

    print(str(datetime.datetime.now()), message1)
    print(str(datetime.datetime.now()), message2)

    return message1, message2


'''
========================
GET LATITUDE & LONGITUDE
========================
'''


def get_coord(loc_auth):
    lat = long = ''

    if pd.isna(loc_auth):
        return lat, long

    '''
    Try to get lat/long from existing data to speed up runtime
    '''

    lat = df_daily[df_daily['areaName'] == loc_auth]['Latitude'].values[0]
    long = df_daily[df_daily['areaName'] == loc_auth]['Longitude'].values[0]

    if lat != '':
        return lat, long

    '''
    Now try to get lat/long from doogal
    '''

    if loc_auth == 'Hackney and City of London':
        loc_auth = 'Hackney'

    if loc_auth == 'Cornwall and Isles of Scilly':
        loc_auth = 'Cornwall'

    if loc_auth == 'Comhairle nan Eilean Siar':
        loc_auth = 'Na h-Eileanan Siar'

    url = 'https://www.doogal.co.uk/AdministrativeAreas.php'

    try:
        source = urllib.request.urlopen(url)
        soup = bs.BeautifulSoup(source, 'lxml')
        tables = soup.find_all('table')

        for tb in tables:
            table_rows = tb.find_all('tr')

            for tr in table_rows:
                thd = tr.find_all(['th', 'td'])
                row = [i.text for i in thd]

                if not row:
                    pass
                else:
                    if row[0] == loc_auth:
                        lat = row[2]
                        long = row[3]
                        break

    except urllib.request.HTTPError as err:
        print('HTTP Error: (local auth ', loc_auth, ')', err.code)

    return lat, long


if __name__ == '__main__':
    app.run_server(debug=True)
